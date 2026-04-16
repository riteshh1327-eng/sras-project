"""
SRAS — Excel Upload Utility
Imports students as permanent identities and creates Enrollment records.

Columns expected: Name | RollID | Email | Gender | DOB
"""

import openpyxl
from datetime import datetime
from .models import Student, Enrollment, StudentClass


EXPECTED_COLUMNS = ['name', 'rollid', 'email', 'gender', 'dob']
VALID_GENDERS    = {'male': 'Male', 'female': 'Female', 'other': 'Other'}


def parse_date(value):
    """Try multiple date formats; return date object or None."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'date'):
        return value.date()
    formats = ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y']
    for fmt in formats:
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def import_students_from_excel(file, student_class: StudentClass) -> dict:
    """
    Parse Excel and create Student + Enrollment records.

    Strategy per row:
      • If a Student with the same email already exists → reuse it
      • Otherwise create a new Student
      • Always create an Enrollment linking the Student to student_class + academic_year
      • Skip if the Enrollment already exists (roll_id already taken in that class+year)

    Returns:
        {created: int, skipped: int, errors: [str], warnings: [str]}
    """
    result = {'created': 0, 'skipped': 0, 'errors': [], 'warnings': []}

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        result['errors'].append(f"Cannot open file: {exc}")
        return result

    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        result['errors'].append("The Excel file is empty.")
        return result

    first_row  = [str(c).strip().lower().replace(' ', '') if c else '' for c in rows[0]]
    has_header = any(col in EXPECTED_COLUMNS for col in first_row)
    data_rows  = rows[1:] if has_header else rows

    if not data_rows:
        result['errors'].append("No data rows found.")
        return result

    ay = student_class.academic_year

    # Pre-fetch existing roll_ids in this class+year to detect duplicates quickly
    existing_rolls = set(
        Enrollment.objects
        .filter(student_class=student_class, academic_year=ay)
        .values_list('roll_id', flat=True)
    )
    seen_rolls_in_batch = set()

    for row_num, row in enumerate(data_rows, start=2 if has_header else 1):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        def col(idx, default=''):
            try:
                v = row[idx]
                return str(v).strip() if v is not None else default
            except IndexError:
                return default

        name       = col(0)
        roll_id    = col(1)
        email      = col(2)
        gender_raw = col(3)
        dob_raw    = row[4] if len(row) > 4 else None

        # Required field validation
        row_errors = []
        if not name:     row_errors.append("Name missing")
        if not roll_id:  row_errors.append("Roll ID missing")
        if not gender_raw: row_errors.append("Gender missing")

        if row_errors:
            result['errors'].append(f"Row {row_num}: {', '.join(row_errors)}")
            result['skipped'] += 1
            continue

        gender = VALID_GENDERS.get(gender_raw.lower())
        if not gender:
            result['warnings'].append(
                f"Row {row_num}: Invalid gender '{gender_raw}' — defaulting to 'Other'"
            )
            gender = 'Other'

        dob = parse_date(dob_raw)
        if dob_raw and not dob:
            result['warnings'].append(
                f"Row {row_num}: Could not parse DOB '{dob_raw}' — left blank"
            )

        # Duplicate roll check
        if roll_id in existing_rolls:
            result['warnings'].append(
                f"Row {row_num}: Roll ID '{roll_id}' already enrolled in {student_class} — skipped"
            )
            result['skipped'] += 1
            continue

        if roll_id in seen_rolls_in_batch:
            result['warnings'].append(
                f"Row {row_num}: Duplicate Roll ID '{roll_id}' in file — skipped"
            )
            result['skipped'] += 1
            continue

        seen_rolls_in_batch.add(roll_id)

        # Get-or-create Student by email (permanent identity)
        student = None
        if email:
            student = Student.objects.filter(email__iexact=email).first()

        if student is None:
            student = Student.objects.create(
                name=name,
                email=email,
                gender=gender,
                date_of_birth=dob,
            )
        else:
            # Update name/DOB if blank
            changed = False
            if not student.date_of_birth and dob:
                student.date_of_birth = dob
                changed = True
            if changed:
                student.save(update_fields=['date_of_birth', 'updated_at'])

        # Create Enrollment
        Enrollment.objects.get_or_create(
            student=student,
            academic_year=ay,
            defaults={
                'student_class': student_class,
                'roll_id':       roll_id,
            },
        )

        existing_rolls.add(roll_id)
        result['created'] += 1

    wb.close()
    return result


def generate_sample_excel():
    """Generate a sample .xlsx for teachers to download."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ['Name', 'RollID', 'Email', 'Gender', 'DOB']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        cell.fill = openpyxl.styles.PatternFill(
            start_color='2563EB', end_color='2563EB', fill_type='solid'
        )

    sample = [
        ['Rahul Patil',   'SE001', 'rahul@example.com',  'Male',   '2004-05-21'],
        ['Priya Sharma',  'SE002', 'priya@example.com',  'Female', '2004-08-15'],
        ['Amit Kumar',    'SE003', 'amit@example.com',   'Male',   '2003-12-10'],
        ['Sneha Desai',   'SE004', 'sneha@example.com',  'Female', '2004-03-25'],
    ]
    for row_data in sample:
        ws.append(row_data)

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    return wb
